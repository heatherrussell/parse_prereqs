from bs4 import element as BSelement
from bs4 import BeautifulSoup
import json

from openpyxl import Workbook
from openpyxl.styles import PatternFill

invalid_tags = ['div', 'span', 'a']
invalid_attributes = ["data-test", "style"]

def strip_html(soup):
    for tag in invalid_tags: 
        for match in soup.findAll(tag):
            match.replaceWithChildren()

    for tag in soup.descendants:
        if isinstance(tag, BSelement.Tag):
            tag.attrs = {key: value for key, value in tag.attrs.items()
                     if key not in invalid_attributes}
    return soup

def parse_contents(li):
    if 'PHYS' in li.contents[0] or 'MATH' in li.contents[0] or 'SEOS' in li.contents[0] or 'CHEM' in li.contents[0] or 'ECE' in li.contents[0] or 'ELEC' in li.contents[0] or 'BME' in li.contents[0]:
        return ' '.join(li.contents[:-1]).strip()
    else: 
        return ''.join(li.find_all(text=True, recursive=False)).strip()

def find_li_better(element):
    result = []
    for ul in element('ul', recursive=False):
        for li in ul('li', recursive=False):   
            contents = parse_contents(li)
            #print(contents)
            nested_list = find_li_better(li)
            if nested_list:
                #print(nested_list)
                result.append({contents:nested_list})  # Add the nested list as a whole
            else:
                #print(contents)
                result.append(contents)  # Add the content directly if there is no nested list
    return result

def postprocess_dict(x):
    if isinstance(x, list) and len(x) == 1:
        x = x[0]
    return x

courses = None
with open('phys_courses.txt', 'r') as file:
    courses = json.load(file)

course_list = [''] #blank at the start so we have the top left cell blank in the matrix

for ic,course in enumerate(courses):
    
    course_list += [course['name']]

    prereq_soup = BeautifulSoup(course['prereqs'].replace('<!-- -->',''), 'html.parser')
    coreq_soup = BeautifulSoup(course['coreqs'].replace('<!-- -->',''), 'html.parser')


    prereq_soup = strip_html(prereq_soup)
    coreq_soup = strip_html(coreq_soup)

    #oh the nesting... this gives the inner lists of courses. need to get to this point!
    #prereq_soup.contents[0].contents[0].contents[0].contents[1].contents[0].contents[0].contents[1].contents[0].contents    
    courses[ic]['prereqs'] = postprocess_dict(find_li_better(prereq_soup))
    courses[ic]['coreqs'] = postprocess_dict(find_li_better(coreq_soup))
   
#import pprint
#pp = pprint.PrettyPrinter(indent=1)
#pp.pprint(courses[22])

wb = Workbook()
ws = wb.active
coreq_fill = PatternFill(start_color="6dd6da", end_color="6dd6da", fill_type="solid")
prereq_fill = PatternFill(start_color="f3bce6", end_color="f3bce6", fill_type="solid")

# write the courses as headers into row1 and col1 
for col_idx, value in enumerate(course_list, 1): 
        cell = ws.cell(row=1, column=col_idx, value=value)
for row_idx, value in enumerate(course_list, 1): 
        cell = ws.cell(row=row_idx, column=1, value=value)

def find_phys_in_dict(data):
    results = []
    if isinstance(data, dict):
        for value in data.values():
            results.extend(find_phys_in_dict(value))
    elif isinstance(data, list):
        for item in data:
            results.extend(find_phys_in_dict(item))
    elif isinstance(data, str):
        if 'PHYS' in data[:4]:
            results.append(data.split('-')[0].strip())
    return results

for ic,course in enumerate(courses): 
        #we'll have it so we can read off pre/co-requisites horizontally
        #need to know what the index of these are...
        #print(course['name'],course['prereqs'])
        for pr in find_phys_in_dict(course['prereqs']):
            try: pr_index = 1+course_list.index(pr)
            except: print("ERROR",pr,'not found in course list')
            cell = ws.cell(row=1+(ic+1), column=pr_index, value='pre')
            cell.fill = prereq_fill
        for pr in find_phys_in_dict(course['coreqs']):
            try: pr_index = 1+course_list.index(pr)
            except: print("ERROR",pr,'not found in course list')
            cell = ws.cell(row=1+(ic+1), column=pr_index, value='co')
            cell.fill = coreq_fill

wb.save("prerequisite_table.xlsx")
